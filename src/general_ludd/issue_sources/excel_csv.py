"""File-based issue source: CSV / Excel (.xlsx) spreadsheets.

System-of-record adapter for tabular issue trackers. Each data row is one
issue. Column mapping is configurable (``id_col``, ``title_col``, ...). Writes
are performed back into the same file (rewrite for CSV, openpyxl save for XLSX).

The ``openpyxl`` import is GUARDED so this module imports cleanly even when
openpyxl is not installed; ``health()`` reports ``openpyxl unavailable`` for an
``.xlsx`` target in that case, and CSV handling works regardless.

Self-contained: standard-library ``csv`` plus optional ``openpyxl``. No base
class, no sibling-module imports.

Path safety: every file path is realpath-confined to the configured ``root``;
``../`` escapes, absolute-outside-root paths, and symlink escapes raise
``ValueError`` at construction time.
"""

from __future__ import annotations

import csv
import os
from types import ModuleType
from typing import Any

try:  # GUARD: module must import fine without openpyxl
    import openpyxl as _openpyxl  # type: ignore[import-untyped]
except ImportError:  # pragma: no cover - environment dependent
    _openpyxl = None

# Re-exported under a stable, monkeypatch-friendly name; None when unavailable.
openpyxl: ModuleType | None = _openpyxl


_DONE_STATUSES = {"done", "closed", "complete", "completed", "resolved", "x"}

_DEFAULT_MAPPING: dict[str, str] = {
    "id_col": "id",
    "title_col": "title",
    "status_col": "status",
    "description_col": "description",
    "assignee_col": "assignee",
    "priority_col": "priority",
    "labels_col": "labels",
}


def _confine(root: str, path: str) -> str:
    """Resolve ``path`` (relative to ``root``) and confine it; raise on escape."""
    root_real = os.path.realpath(root)
    candidate = path if os.path.isabs(path) else os.path.join(root_real, path)
    target_real = os.path.realpath(candidate)
    if target_real != root_real and not target_real.startswith(root_real + os.sep):
        raise ValueError(f"path {path!r} escapes the configured root {root!r}")
    return target_real


class ExcelCsvSource:
    """Issue source backed by a CSV or XLSX spreadsheet."""

    SYSTEM = "excel"

    def __init__(self, config: dict[str, Any]) -> None:
        root = config.get("root")
        path = config.get("path")
        if not root:
            raise ValueError("config['root'] (allowed base directory) is required")
        if not path:
            raise ValueError("config['path'] is required")
        self._root = os.path.realpath(str(root))
        self._path = _confine(self._root, str(path))
        self._mapping: dict[str, str] = dict(_DEFAULT_MAPPING)
        for key in _DEFAULT_MAPPING:
            if config.get(key):
                self._mapping[key] = str(config[key])
        self.name = f"excel:{os.path.relpath(self._path, self._root)}"

    # -- internals --------------------------------------------------------

    def _rel(self) -> str:
        return os.path.relpath(self._path, self._root)

    def _is_xlsx(self) -> bool:
        return self._path.lower().endswith(".xlsx")

    def _file_mtime(self) -> float | None:
        try:
            return os.path.getmtime(self._path)
        except OSError:
            return None

    def _read_rows(self) -> tuple[list[str], list[dict[str, str]]]:
        """Return ``(header, rows)`` where rows are header->value dicts."""
        if self._is_xlsx():
            if openpyxl is None:
                raise RuntimeError("openpyxl unavailable: cannot read .xlsx files")
            wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
            ws = wb.active
            values = list(ws.iter_rows(values_only=True))
            wb.close()
            if not values:
                return [], []
            header = [("" if c is None else str(c)) for c in values[0]]
            rows: list[dict[str, str]] = []
            for raw in values[1:]:
                if raw is None or all(c is None for c in raw):
                    continue
                cells = ["" if c is None else str(c) for c in raw]
                rows.append({header[i]: (cells[i] if i < len(cells) else "") for i in range(len(header))})
            return header, rows
        with open(self._path, newline="", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            header = list(reader.fieldnames or [])
            rows = [{k: ("" if v is None else str(v)) for k, v in row.items()} for row in reader]
        return header, rows

    def _normalize(self, row: dict[str, str], index: int, mtime: float | None) -> dict[str, Any]:
        mp = self._mapping
        ext = row.get(mp["id_col"], "").strip() or f"row-{index + 1}"
        raw_status = row.get(mp["status_col"], "").strip()
        status = "done" if raw_status.lower() in _DONE_STATUSES else "open"
        labels_raw = row.get(mp["labels_col"], "")
        labels = [tok.strip() for tok in labels_raw.split(",") if tok.strip()] if labels_raw else []
        return {
            "external_id": ext,
            "source": self.SYSTEM,
            "title": row.get(mp["title_col"], "").strip(),
            "description": row.get(mp["description_col"], "").strip(),
            "status": status,
            "assignee": (row.get(mp["assignee_col"], "").strip() or None),
            "labels": labels,
            "priority": (row.get(mp["priority_col"], "").strip() or None),
            "url": f"{self._path}#row={index + 1}",
            "updated_ts": mtime,
            "raw": dict(row),
        }

    def _write_rows(self, header: list[str], rows: list[dict[str, str]]) -> None:
        if self._is_xlsx():
            if openpyxl is None:
                raise RuntimeError("openpyxl unavailable: cannot write .xlsx files")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(header)
            for row in rows:
                ws.append([row.get(col, "") for col in header])
            wb.save(self._path)
            return
        with open(self._path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=header)
            writer.writeheader()
            for row in rows:
                writer.writerow({col: row.get(col, "") for col in header})

    # -- contract ---------------------------------------------------------

    def health(self) -> dict[str, Any]:
        if self._is_xlsx() and openpyxl is None:
            return {"ok": False, "detail": "openpyxl unavailable: cannot read .xlsx"}
        try:
            with open(self._path, "rb"):
                pass
        except OSError as exc:
            return {"ok": False, "detail": f"unreadable: {exc}"}
        return {"ok": True, "detail": f"readable: {self._rel()}"}

    def fetch_issues(self, spec: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        _, rows = self._read_rows()
        mtime = self._file_mtime()
        return [self._normalize(row, idx, mtime) for idx, row in enumerate(rows)]

    def update_status(self, external_id: str, status: str, comment: str | None = None) -> dict[str, Any]:
        header, rows = self._read_rows()
        id_col = self._mapping["id_col"]
        status_col = self._mapping["status_col"]
        matched = False
        is_done = status.strip().lower() in _DONE_STATUSES
        new_value = "done" if is_done else "open"
        for row in rows:
            if row.get(id_col, "").strip() == external_id:
                row[status_col] = new_value
                if comment and "comment" in header:
                    row["comment"] = comment
                matched = True
        if not matched:
            raise KeyError(f"external_id {external_id!r} not found in {self._rel()}")
        if status_col not in header:
            header = [*header, status_col]
        self._write_rows(header, rows)
        return {
            "external_id": external_id,
            "status": new_value,
            "source": self.SYSTEM,
            "url": f"{self._path}",
            "comment": comment,
        }

    def add_comment(self, external_id: str, comment: str) -> dict[str, Any]:
        header, rows = self._read_rows()
        id_col = self._mapping["id_col"]
        matched = False
        for row in rows:
            if row.get(id_col, "").strip() == external_id:
                existing = row.get("comment", "")
                row["comment"] = f"{existing}\n{comment}".strip() if existing else comment
                matched = True
        if not matched:
            raise KeyError(f"external_id {external_id!r} not found in {self._rel()}")
        if "comment" not in header:
            header = [*header, "comment"]
        self._write_rows(header, rows)
        return {
            "external_id": external_id,
            "source": self.SYSTEM,
            "comment": comment,
            "url": f"{self._path}",
        }
